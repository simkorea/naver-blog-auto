# -*- coding: utf-8 -*-
"""
증거목록표 (xlsx) — 변호사가 그대로 쓸 수 있는 형태.

담기는 것
  번호 · 파일명 · 종류 · 발생일시 · 위치(타임코드/쪽) · 화자 ·
  발언 내용 · 쟁점 · 유불리 · 전사 신뢰도 · 청취 확인 · 제출 사유 · SHA-256

신뢰도와 청취 확인을 굳이 표에 넣는 이유
  AI 전사는 틀릴 수 있다. 확인하지 않은 구간을 확정 사실처럼 제출했다가
  나중에 오인식이 드러나면, 그 하나 때문에 나머지 증거의 신빙성까지
  통째로 의심받는다. 미검증은 미검증이라고 적어두는 편이 훨씬 안전하다.
"""
from datetime import datetime
from pathlib import Path

FONT = "Arial"

HEADERS = [
    ("번호", 6), ("파일명", 30), ("종류", 8), ("발생일시", 17),
    ("위치", 20), ("화자", 12), ("발언·내용", 60), ("쟁점", 18),
    ("유불리", 8), ("전사 신뢰도", 16), ("확인", 11),
    ("제출 사유", 28), ("SHA-256 (원본)", 66),
]

KIND_KR = {"audio": "녹음", "kakao": "카톡", "document": "문서",
           "image": "이미지", "email": "메일"}

# 엑셀 셀 한 칸에 들어갈 수 있는 최대 글자 수.
# 넘으면 엑셀이 파일을 열지 못하거나 내용을 잘라버린다.
CELL_LIMIT = 32000


def _fit(value):
    """셀 한도를 넘으면 잘라내되, 잘렸다는 사실을 남긴다."""
    if isinstance(value, str) and len(value) > CELL_LIMIT:
        return value[:CELL_LIMIT] + " …(이하 생략)"
    return value


def write(rows: list[dict], out_path, title: str = "증거목록") -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    out_path = Path(out_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "증거목록"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 머리말 ────────────────────────────────
    ws["A1"] = title
    ws["A1"].font = Font(name=FONT, size=15, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    ws["A2"] = (f"작성 {datetime.now():%Y년 %m월 %d일 %H:%M}　·　총 {len(rows)}건　·　"
                "SHA-256은 수집 시점에 기록한 원본의 해시값입니다")
    ws["A2"].font = Font(name=FONT, size=9, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))

    ws["A3"] = ("※ '확인' 열이 '미검증'인 항목은 AI 자동 전사 결과를 아직 원본 음성과 "
                "대조하지 않은 구간입니다. 제출 전 확인이 필요합니다.")
    ws["A3"].font = Font(name=FONT, size=9, color="C00000")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(HEADERS))

    # ── 열 머리 ───────────────────────────────
    HEAD_ROW = 5
    head_fill = PatternFill("solid", fgColor="1F3864")
    for c, (name, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=HEAD_ROW, column=c, value=name)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[HEAD_ROW].height = 26

    # ── 본문 ─────────────────────────────────
    good_fill = PatternFill("solid", fgColor="E2EFDA")     # 유리
    bad_fill = PatternFill("solid", fgColor="FCE4E4")      # 불리
    warn_font = Font(name=FONT, size=10, color="C00000", bold=True)
    base_font = Font(name=FONT, size=10)

    for i, r in enumerate(rows):
        row = HEAD_ROW + 1 + i
        values = [
            r["no"], r["file"], KIND_KR.get(r["kind"], r["kind"]), r["when"],
            r["location"], r["speaker"], r["text"], r["issue"],
            r["stance"], r["reliability"], r["verified"],
            r["reason"], r["sha256"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=_fit(v))
            cell.font = base_font
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(c in (2, 7, 8, 12)),
                horizontal="center" if c in (1, 3, 5, 9, 10, 11) else "left",
            )

        if r["stance"] == "유리":
            ws.cell(row=row, column=9).fill = good_fill
        elif r["stance"] == "불리":
            ws.cell(row=row, column=9).fill = bad_fill

        # 확인·신뢰도 경고는 눈에 띄어야 한다
        if r["verified"] == "미검증":
            ws.cell(row=row, column=11).font = warn_font
        if "확인 필요" in r["reliability"] or r["reliability"] == "낮음":
            ws.cell(row=row, column=10).font = warn_font

        ws.row_dimensions[row].height = 30

    ws.freeze_panes = ws.cell(row=HEAD_ROW + 1, column=1)
    ws.auto_filter.ref = (f"A{HEAD_ROW}:"
                          f"{get_column_letter(len(HEADERS))}{HEAD_ROW + len(rows)}")

    _summary_sheet(wb, rows)
    wb.save(out_path)
    return out_path


def _summary_sheet(wb, rows):
    """
    쟁점별 집계. 어디에 증거가 몰려 있고 어디가 비었는지 보인다.

    숫자를 박아 넣지 않고 증거목록 시트를 COUNTIFS로 세는 이유:
    변호사가 목록에서 항목을 지우거나 유불리를 고치면 집계가 따라 바뀐다.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet("쟁점별 집계")
    ws["A1"] = "쟁점별 증거 집계"
    ws["A1"].font = Font(name=FONT, size=14, bold=True)
    ws["A2"] = "증거목록 시트를 실시간으로 집계합니다. 목록을 고치면 이 숫자도 바뀝니다."
    ws["A2"].font = Font(name=FONT, size=9, color="666666")

    # 증거목록 시트에서 세어야 할 범위 (머리행 다음부터)
    first, last = 6, 5 + max(len(rows), 1)
    ISSUE_COL, STANCE_COL, VERIFY_COL = "H", "I", "K"
    src = "증거목록"

    issues = []
    for r in rows:
        for issue in [x.strip() for x in (r["issue"] or "").split(",") if x.strip()]:
            if issue not in issues:
                issues.append(issue)

    head = ["쟁점", "유리", "불리", "중립", "미검증", "합계"]
    for c, name in enumerate(head, 1):
        cell = ws.cell(row=4, column=c, value=name)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 26
    for col in "BCDEF":
        ws.column_dimensions[col].width = 10

    for i, issue in enumerate(issues):
        row = 5 + i
        ws.cell(row=row, column=1, value=issue).font = Font(name=FONT, size=10)
        rng_issue = f"{src}!${ISSUE_COL}${first}:${ISSUE_COL}${last}"
        rng_stance = f"{src}!${STANCE_COL}${first}:${STANCE_COL}${last}"
        rng_verify = f"{src}!${VERIFY_COL}${first}:${VERIFY_COL}${last}"

        for c, stance in enumerate(("유리", "불리", "중립"), 2):
            # 쟁점 열은 쉼표로 여러 쟁점이 들어갈 수 있어 부분 일치로 센다
            cell = ws.cell(
                row=row, column=c,
                value=f'=COUNTIFS({rng_issue},"*"&$A{row}&"*",{rng_stance},"{stance}")')
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(horizontal="center")

        cell = ws.cell(
            row=row, column=5,
            value=f'=COUNTIFS({rng_issue},"*"&$A{row}&"*",{rng_verify},"미검증")')
        cell.font = Font(name=FONT, size=10, color="C00000")
        cell.alignment = Alignment(horizontal="center")

        total = ws.cell(row=row, column=6, value=f"=SUM(B{row}:D{row})")
        total.font = Font(name=FONT, size=10, bold=True)
        total.alignment = Alignment(horizontal="center")

    if issues:
        last_row = 4 + len(issues)
        ws.cell(row=last_row + 1, column=1, value="합계").font = Font(
            name=FONT, size=10, bold=True)
        for c in range(2, 7):
            col = chr(ord("A") + c - 1)
            cell = ws.cell(row=last_row + 1, column=c,
                           value=f"=SUM({col}5:{col}{last_row})")
            cell.font = Font(name=FONT, size=10, bold=True)
            cell.alignment = Alignment(horizontal="center")
